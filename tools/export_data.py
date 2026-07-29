# -*- coding: utf-8 -*-
"""
export_data.py — експорт даних лонгріда з оновленого Excel-файлу
«Бюджети_участі_великих_міст_України_2015-2026_міжнародне_порівняння_*.xlsx».

Пише:
  data/pb_data.json    — 35 міст × 2015–2026: показники, статуси, річні агрегати
  data/analytics.json  — частка БУ у видатках (131 спостереження), профілі міст
                         і міжнародні кейси

Запуск: python tools/export_data.py  (з папки interactive-pb-article/ або кореня)
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    PROJECT, '..', 'Аналітика Бюджетів участі',
    'Бюджети_участі_великих_міст_України_2015-2026_міжнародне_порівняння_2026-07-23.xlsx')

YEARS = list(range(2015, 2027))

# Координати міст (WGS84: lon, lat)
CITY_COORDS = {
    'Бердянськ': (36.799, 46.756), 'Біла Церква': (30.117, 49.796), 'Бровари': (30.790, 50.511),
    'Вінниця': (28.468, 49.233), 'Дніпро': (35.046, 48.464), 'Житомир': (28.658, 50.254),
    'Запоріжжя': (35.139, 47.838), 'Івано-Франківськ': (24.711, 48.922), "Кам'янське": (34.602, 48.511),
    'Київ': (30.523, 50.450), 'Краматорськ': (37.584, 48.739), 'Кременчук': (33.420, 49.068),
    'Кривий Ріг': (33.392, 47.910), 'Кропивницький': (32.262, 48.507), 'Луцьк': (25.325, 50.747),
    'Львів': (24.031, 49.842), 'Маріуполь': (37.543, 47.097), 'Мелітополь': (35.365, 46.848),
    'Миколаїв': (31.995, 46.975), 'Нікополь': (34.396, 47.567), 'Одеса': (30.723, 46.482),
    'Павлоград': (35.870, 48.520), 'Полтава': (34.551, 49.589), 'Рівне': (26.251, 50.620),
    'Сєвєродонецьк': (38.492, 48.948), "Слов'янськ": (37.598, 48.854), 'Суми': (34.798, 50.907),
    'Тернопіль': (25.594, 49.553), 'Ужгород': (22.288, 48.621), 'Харків': (36.231, 49.994),
    'Херсон': (32.617, 46.635), 'Хмельницький': (26.987, 49.423), 'Черкаси': (32.060, 49.444),
    'Чернівці': (25.935, 48.292), 'Чернігів': (31.294, 51.494),
}

REGION_ISO = {
    'Вінницька область': 'UA-05', 'Волинська область': 'UA-07', 'Луганська область': 'UA-09',
    'Дніпропетровська область': 'UA-12', 'Донецька область': 'UA-14', 'Житомирська область': 'UA-18',
    'Закарпатська область': 'UA-21', 'Запорізька область': 'UA-23', 'Івано-Франківська область': 'UA-26',
    'м. Київ': 'UA-30', 'Київська область': 'UA-32', 'Кіровоградська область': 'UA-35',
    'Львівська область': 'UA-46', 'Миколаївська область': 'UA-48', 'Одеська область': 'UA-51',
    'Полтавська область': 'UA-53', 'Рівненська область': 'UA-56', 'Сумська область': 'UA-59',
    'Тернопільська область': 'UA-61', 'Харківська область': 'UA-63', 'Херсонська область': 'UA-65',
    'Хмельницька область': 'UA-68', 'Черкаська область': 'UA-71', 'Чернігівська область': 'UA-74',
    'Чернівецька область': 'UA-77',
}

norm = lambda s: str(s).replace('’', "'").replace('ʼ', "'").strip()

# Частина рядків має код замість української назви основи суми — перекладаємо
BASIS_UA = {
    'annual_envelope_approx': 'річний програмний ліміт (приблизно)',
    'implementation_plan': 'план реалізації проєктів',
    'planned_annual_envelope_not_actual': 'плановий річний ліміт (без підтвердженого факту)',
    'program_forecast_not_actual': 'прогноз програми (без підтвердженого факту)',
    'program_setup_amount_not_project_envelope': 'сума на запровадження програми',
    'reported_winner_total': 'повідомлена сумарна вартість переможців',
}


def basis_ua(v):
    if v is None:
        return None
    v = norm(v)
    return BASIS_UA.get(v, v)


def fnum(v):
    if v is None or v == '':
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------------------------------------------------------- Дані_місто-рік
ws = wb['Дані_місто-рік']
rows = [r for r in ws.iter_rows(values_only=True)]
header_idx = next(i for i, r in enumerate(rows) if r[0] == 'Місто' and r[1])
records = [r for r in rows[header_idx + 1:] if r[0]]

cities = {}
for r in records:
    name = norm(r[0])
    year = int(r[9])
    if year not in YEARS:
        continue
    if name not in cities:
        lon, lat = CITY_COORDS[name]
        region = norm(r[1])
        cities[name] = {
            'name': name, 'region': region, 'iso': REGION_ISO[region],
            'lon': lon, 'lat': lat,
            'population': int(fnum(r[2]) or 0),
            'tot2022': bool(r[4] and 'ТОТ' in str(r[4])),
            'firstYear': int(fnum(r[7])) if fnum(r[7]) else None,
            'years': {},
        }
    a, w, s = fnum(r[12]), fnum(r[13]), fnum(r[14])
    budget = fnum(r[18])
    completeness = str(r[23] or '').strip()
    per_winner = fnum(r[31])

    # Стан клітинки для карти: порожньо ≠ нуль, стани розрізняються явно
    has_metric = any(v is not None for v in (a, w, s))
    if has_metric:
        status = 'data'
    elif completeness == 'not_applicable':
        status = 'pre'          # до запуску програми
    elif completeness == 'suspended':
        status = 'susp'         # цикл офіційно призупинено
    elif completeness == 'not_held':
        status = 'notheld'      # конкурс не проводився
    else:                       # missing / context_only / cycle_confirmed_no_metrics
        status = 'nodata'

    cities[name]['years'][year] = {
        'a': int(a) if a is not None else None,
        'w': int(w) if w is not None else None,
        's': s,
        'budget': budget,
        'status': status,
        'basis': basis_ua(r[16]),
        'precision': str(r[17] or '') or None,
        'perWinner': round(per_winner) if per_winner else None,
        'share': None,  # заповнюється нижче лише для зіставних спостережень
    }

assert len(cities) == 35, f'Очікували 35 міст, отримали {len(cities)}'

# ------------------------------------------------- Спостереження_частки (131)
ws = wb['Спостереження_частки']
obs_rows = [r for r in ws.iter_rows(values_only=True)]
obs_header = next(i for i, r in enumerate(obs_rows) if r[0] == 'Місто')
observations = []
for r in obs_rows[obs_header + 1:]:
    if not r[0]:
        continue
    name, year = norm(r[0]), int(r[1])
    share = fnum(r[4])
    observations.append({
        'city': name, 'year': year,
        's': fnum(r[2]), 'budget': fnum(r[3]), 'share': share,
        'basis': basis_ua(r[6]),
        'precision': str(r[7] or '') or None,
    })
    if name in cities and year in cities[name]['years']:
        cities[name]['years'][year]['share'] = share

assert len(observations) == 131, f'Очікували 131 спостереження, отримали {len(observations)}'

# ---------------------------------------------------------- річні агрегати
totals = {}
for y in YEARS:
    active = apps = wins = amount = with_amount = 0
    for c in cities.values():
        d = c['years'].get(y)
        if not d:
            continue
        if d['status'] == 'data':
            active += 1
        if d['a'] is not None:
            apps += d['a']
        if d['w'] is not None:
            wins += d['w']
        if d['s'] is not None:
            amount += d['s']
            with_amount += 1
    totals[y] = {
        'activeCities': active, 'applications': apps, 'winners': wins,
        'amount': round(amount), 'citiesWithAmount': with_amount,
        'confirmedOnly': y >= 2022,  # після 2022 — документований мінімум
    }

city_totals = sorted(
    ({'name': c['name'],
      'amount': round(sum(d['s'] or 0 for d in c['years'].values()))}
     for c in cities.values()),
    key=lambda x: -x['amount'])

# ---------------------------------------------------------- Аналітика_частки
ws = wb['Аналітика_частки']
an = [list(r) for r in ws.iter_rows(values_only=True)]
sum_hdr = next(i for i, r in enumerate(an) if str(r[0] or '').startswith('Середнє місто–рік'))
sum_vals = [fnum(c) for c in an[sum_hdr + 1] if fnum(c) is not None]
summary = {
    'mean': sum_vals[0], 'median': sum_vals[1],
    'weighted': sum_vals[2], 'n': int(sum_vals[3]),
    'frame': '35 міст, 2018–2026',
}
by_year = []
for r in an:
    y = fnum(r[0]) if r and r[0] is not None else None
    if y and 2018 <= y <= 2026 and fnum(r[2]) is not None and fnum(r[4]) is not None:
        by_year.append({'year': int(y), 'n': int(fnum(r[1])),
                        'mean': fnum(r[2]), 'median': fnum(r[3]), 'weighted': fnum(r[4])})
by_year = sorted({d['year']: d for d in by_year}.values(), key=lambda d: d['year'])

# профілі міст (ранжовані за середньою часткою) — блок після рядка-заголовка
profiles = []
prof_start = next(i for i, r in enumerate(an) if r[0] == 'Місто' and r[1] == 'N')
for r in an[prof_start + 1:]:
    if not r[0] or fnum(r[4]) is None:
        break
    profiles.append({
        'city': norm(r[0]), 'n': int(fnum(r[1])),
        'firstYear': int(fnum(r[2])), 'lastYear': int(fnum(r[3])),
        'mean': fnum(r[4]), 'median': fnum(r[5]),
        'min': fnum(r[7]), 'max': fnum(r[8]),
    })

periods = [
    {'label': '2018–2021', 'n': 111, 'mean': 0.0030754003549992447},
    {'label': '2022–2026', 'n': 20, 'mean': 0.002941852073873755},
]
for r in an:  # блок «Періоди» може бути в довільних колонках праворуч
    for ci, cell in enumerate(r):
        lab = str(cell or '').strip()
        if lab in ('2018–2021', '2022–2026'):
            tail = [fnum(c) for c in r[ci + 1:] if fnum(c) is not None]
            if len(tail) >= 2:
                idx = 0 if lab == '2018–2021' else 1
                periods[idx] = {'label': lab, 'n': int(tail[0]), 'mean': tail[1]}

# Старий аркуш «Аналітичний_поріг» містить авторські сценарні орієнтири,
# зокрема помилково названий «інституційний мінімум» 0,5 %. Вони не є
# нормативами, не використовуються інтерфейсом і тому не експортуються.
thresholds = []

# ---------------------------------------------------------- Норми_порівняння
ws = wb['Норми_порівняння']
intl = []
for r in ws.iter_rows(values_only=True):
    if not r[0] or r[0] in ('Географія',) or 'Правові норми' in str(r[0]):
        continue
    if r[2] is None or not str(r[10] or '').startswith('http'):
        continue
    intl.append({
        'place': str(r[0]).strip(), 'period': str(r[1] or '').strip(),
        'case': str(r[2]).strip(), 'sourceType': str(r[3] or '').strip(),
        'status': str(r[4] or '').strip(), 'limit': str(r[5] or '').strip(),
        'denominator': str(r[6] or '').strip(), 'shareLabel': str(r[7] or '').strip(),
        'claim': str(r[8] or '').strip(),
        'interpretation': str(r[9] or '').strip(),
    })

# ---------------------------------------------------------------------- запис
pb = {
    'meta': {
        'title': 'Бюджети участі великих міст України, 2015–2026',
        'generated': '2026-07-23',
        'frame': ('35 міст: 34 міста з населенням понад 100 тис. осіб (на 01.01.2022) '
                  'та Сєвєродонецьк як обґрунтований виняток'),
        'caveats': [
            'Порожня клітинка не дорівнює нулю: стани «немає даних», «до запуску», '
            '«призупинено», «не проводився» розрізняються явно.',
            'Суми мають різні основи (річний ліміт, сума бюджетів переможців, фактичні видатки) '
            'і не усереднюються між містами.',
            'Показники 2022–2026 років — документований мінімум за підтвердженими циклами.',
        ],
    },
    'years': YEARS,
    'totals': {str(y): totals[y] for y in YEARS},
    'cities': sorted(cities.values(), key=lambda c: c['name']),
    'cityTotals': [c for c in city_totals if c['amount'] > 0],
}

analytics = {
    'summary': summary,
    'byYear': by_year,
    'observations': observations,
    'profiles': profiles,
    'periods': periods,
    'thresholds': thresholds,
    'international': intl,
    'barcelona': {'programEur': 30000000, 'yearsSpan': 4, 'annualEur': 7500000,
                  'cityBudgetEur': 3807000000, 'share': 0.001970055161544523},
}

os.makedirs(os.path.join(PROJECT, 'data'), exist_ok=True)
with open(os.path.join(PROJECT, 'data', 'pb_data.json'), 'w', encoding='utf-8') as f:
    json.dump(pb, f, ensure_ascii=False)
with open(os.path.join(PROJECT, 'data', 'analytics.json'), 'w', encoding='utf-8') as f:
    json.dump(analytics, f, ensure_ascii=False)

# ------------------------------------------------------------------- контроль
print('Міст:', len(cities), '| спостережень частки:', len(observations),
      '| профілів:', len(profiles), '| міжнародних рядків:', len(intl),
      '| порогів:', len(thresholds))
print('Середня частка: {:.4%} | медіана: {:.4%} | зважена: {:.4%} | n={}'.format(
    summary['mean'], summary['median'], summary['weighted'], summary['n']))
for y in YEARS:
    t = totals[y]
    print(f"{y}: активних {t['activeCities']}, заявок {t['applications']}, "
          f"переможців {t['winners']}, сума {t['amount']/1e6:.1f} млн")
print('ТОП-6 за сумою 2015–2026:',
      ', '.join(f"{c['name']} {c['amount']/1e6:.0f} млн" for c in city_totals[:6]))
print('Роки частки:', [d['year'] for d in by_year])
